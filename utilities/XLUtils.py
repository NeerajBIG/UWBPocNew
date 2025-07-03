import os, sys
from os.path import dirname, join, abspath
sys.path.insert(0, abspath(join(dirname(__file__), '..')))
import openpyxl

def readDataConfig(file, sheetName, ToFind):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    for r in range(1, rows + 1):
        for c in range(1, 2):
            cellData = sheet.cell(row=r, column=c).value
            if cellData == ToFind:
                return sheet.cell(row=r, column=c + 1).value
                break

def readDataTestUserData(file, sheetName, ToFind):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    for r in range(1, rows + 1):
        # print("r is " + str(r))
        for c in range(1, 2):
            cellData = sheet.cell(row=r, column=c).value
            if cellData == ToFind:
                return sheet.cell(row=r, column=c + 1).value
                break

def readDataTestUserDataConfirmation(file, sheetName, ToFind):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    for r in range(1, rows + 1):
        for c in range(1, 2):
            cellData = sheet.cell(row=r, column=c).value
            if cellData == ToFind:
                return sheet.cell(row=r, column=c + 2).value
                break

def readXLData(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    columns = sheet.max_column
    dataList = {}
    for r in range(1, rows + 1):
        for c1 in range(1, columns + 1):
            data = sheet.cell(row=r, column=c1).value
            if data is None or data == "None":
                pass
            else:
                dataValue = sheet.cell(row=r, column=c1+1).value
                dataList[data]=dataValue
    return dataList

def getAllScenarios(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    columns = sheet.max_column
    dataList1 = {}
    for r in range(2, rows + 1):
        for c1 in range(1, 3):
            data = sheet.cell(row=r, column=c1).value
            if data is None or data == "None":
                pass
            else:
                dataValue1 = sheet.cell(row=r, column=c1+1).value
                dataList1[data]=dataValue1
    return dataList1

def readDataScenarios(file, sheetName, ToFind):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    columns = sheet.max_column
    dataList = []
    for r in range(2, rows + 1):
        for c in range(1, 2):
            Scenario_ID = sheet.cell(row=r, column=c).value
            if Scenario_ID == ToFind:
                for c1 in range(1, columns + 1):
                    data = sheet.cell(row=r, column=c1).value
                    if data is None or data == "None":
                        pass
                    else:
                        if c1 > 4:
                            data = data[data.find("$-") + len("$-"):data.rfind("-$")]
                        dataList.append(data)
                return dataList


def writeDataReport(file, sheetName, TestScenarioID, TestScenarioTitle, TestStep, Error, StartTime, EndTime):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    rows = sheet.max_row
    columns = sheet.max_column
    for r in range(rows + 1, rows + 2):
        for c in range(1, columns + 1):
            cellData = sheet.cell(row=r, column=c).value
            #print(cellData)
            if cellData is None or cellData == "None":
                if c == 1:
                    sheet.cell(row=r, column=c).value = TestScenarioID
                if c == 2:
                    sheet.cell(row=r, column=c).value = TestScenarioTitle
                if c == 3:
                    sheet.cell(row=r, column=c).value = TestStep
                if c == 4:
                    sheet.cell(row=r, column=c).value = Error
                if c == 5:
                    sheet.cell(row=r, column=c).value = StartTime
                if c == 6:
                    sheet.cell(row=r, column=c).value = EndTime
                if c == 7:
                    time_difference = EndTime - StartTime
                    sheet.cell(row=r, column=c).value = round(float(time_difference.total_seconds()), 2)
                if c == 8:
                    time_difference = EndTime - StartTime
                    sheet.cell(row=r, column=c).value = round(float((time_difference.total_seconds()) / 60), 3)
        workbook.save(file)
        break
