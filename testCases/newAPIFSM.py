import datetime
import json
import time
import openpyxl
from utilities.readProperties import ReadConfig
import requests

#--General configurations
base_url_ProdMirror = 'https://trires-triliv-prod1-mirror.pegacloud.net/prweb/'
base_url_Stage = 'https://trires-triliv-stg1.pegacloud.net/prweb/'

base_url = base_url_ProdMirror

#--Setting CLIENT_ID and CLIENT_SECRET
CLIENT_SECRET = "none"
CLIENT_ID = "none"
if "stg" in base_url:
    print("It is stg url")
    CLIENT_ID = "63938493146423881858"
    CLIENT_SECRET = "EC723A1823B4D33C6594C3E25CAC91D7"

elif "prod1-mirror" in base_url:
    print("It is prod1-mirror url")
    CLIENT_ID = "63938493146423881858"
    CLIENT_SECRET = "CF71D21FBDF9540A05703460F5855AEA"

grant_type = 'client_credentials'
body_params = {'grant_type': grant_type}

#--Reading API Data excel file
basePath = ReadConfig.basePath()
APIDataSheetName = "APIData.xlsx"
dataSheetPath = basePath + "/TestData/" + APIDataSheetName
sheetName_Config = "ResidentsInfo"
workbook = openpyxl.load_workbook(dataSheetPath)
sheet = workbook[sheetName_Config]
rows = sheet.max_row
columns = sheet.max_column

#--reading Resident details from API Data excel file
ResidentDataList = []
ResidentDataDict = {}
for r in range(2, rows + 1):
    ResidentDataList.clear()
    for c in range(1, columns + 1):
        Data = sheet.cell(row=r, column=c).value
        ResidentDataList.append(Data)
    ResidentDataDict[ResidentDataList[0]] = ResidentDataList[1:5]
#print(ResidentDataDict)

#--Running loop for WO creation based on the Residents count
WOCounter = 0
for ResidentName, value in ResidentDataDict.items():
    print("")
    print("")
    print("*************************Running for Resident - " + ResidentName, value)

    dict_option_queryID = {}

    #--Looping different issue types, based on the count value. max count number is 26:
    count = 1
    for pyIDCount in range(0, count):
        print(" ")
        start = time.time()
        sheetName_Config = "PyIDList"
        workbook = openpyxl.load_workbook(dataSheetPath)
        sheet = workbook[sheetName_Config]
        rows = sheet.max_row
        columns = sheet.max_column

        #--Generating fresh token before each request
        AuthUrl = base_url + 'PRRestService/oauth2/v1/token'
        response = requests.post(AuthUrl, data=body_params, auth=(CLIENT_ID, CLIENT_SECRET))
        data_raw = json.loads(response.text)
        token = data_raw["access_token"]
        #print(token)

        #=====================================================================Get request 1 to get list of options with QueryID=====================================================================G
        headers = {"Authorization": "Bearer {}".format(token)}
        getD1API = '/v1/data/D_DecisionTree?DecisionID=D-1'
        url = base_url + "api/TriconServices"
        r = requests.get(url=url + getD1API, headers=headers)
        #print(r.text)
        d1 = json.loads(r.text)

        pyID = d1["pyID"]
        try:
            optionSelected = d1["pxResults"][pyIDCount]["Option"]

            dict_option_queryID[d1["pxResults"][pyIDCount]["Option"]] = d1["pxResults"][pyIDCount]["QueryID"]
            optionQuery = dict_option_queryID[d1["pxResults"][pyIDCount]["Option"]]
            #print(optionQuery)

            # ----------------Inserting data in Excel
            DataList = [pyID, "Unused", optionSelected]
            #columns = 100
            DataListCounter = 0
            for c in range(1, columns + 1):
                cellData = sheet.cell(row=pyIDCount + 2, column=c).value
                if cellData is None or cellData == "None":
                    sheet.cell(row=pyIDCount + 2, column=c).value = DataList[DataListCounter]
                    DataListCounter = DataListCounter + 1
                    if DataListCounter == len(DataList):
                        workbook.save(dataSheetPath)
                        break

            if "&" in optionSelected:
                optionSelected = optionSelected.replace("&", "%26")
                #print(optionSelected)

            #=====================================================================Get request 2 to get question set 2=====================================================================G
            headers = {"Authorization": "Bearer {}".format(token)}
            getD2API = '/v1/data/D_DecisionTree?DecisionID=' + 'D-1' + '&QuestionID=' + optionQuery + '&Response=' + optionSelected + '&RunID=' + pyID
            #print("getD2API - " + getD2API)
            r2 = requests.get(url=url + getD2API, headers=headers)
            #print(r2.text)
            d2 = json.loads(r2.text)
            Option2 = d2["pxResults"][0]["Option"]
            Option2QueryID = d2["pxResults"][0]["QueryID"]
            Option2DiagnosticID = d2["question"]["DiagnosticID"]
            Option2Question = d2["question"]["Question"]

            # ----------------Inserting data in Excel
            DataList = [Option2Question, Option2]
            columns = 100
            DataListCounter = 0
            for c in range(1, columns + 1):
                cellData = sheet.cell(row=pyIDCount + 2, column=c).value
                if cellData is None or cellData == "None":
                    sheet.cell(row=pyIDCount + 2, column=c).value = DataList[DataListCounter]
                    DataListCounter = DataListCounter + 1
                    if DataListCounter == len(DataList):
                        workbook.save(dataSheetPath)
                        break

            OptionN = "N/A"
            OptionNQueryID = "N/A"
            OptionNDiagnosticID = "N/A"
            OptionNQuestion = "N/A"
            API_count = 10
            for APICount in range(1, API_count):
                if APICount == 1:
                    OptionN = Option2
                    OptionNQueryID = Option2QueryID
                    OptionNDiagnosticID = Option2DiagnosticID
                    OptionNQuestion = Option2Question
                #=====================================================================Get request 3 to get question set 3=====================================================================
                headers = {"Authorization": "Bearer {}".format(token)}
                getD3API = '/v1/data/D_DecisionTree?DecisionID=' + OptionNDiagnosticID + '&QuestionID=' + OptionNQueryID + '&Response=' + OptionN + '&RunID=' + pyID
                #print("getD3API - " + getD3API)
                rN = requests.get(url=url + getD3API, headers=headers)
                #print(rN.text)

                try:
                    dN = json.loads(rN.text)
                    isLastQuestion = dN["question"]["Question"]

                    if "Do you have any additional comments?" in isLastQuestion or "Do you have any comments regarding the issue you are reporting?" in isLastQuestion:
                        OptionN = "This is a test comment"
                        OptionNQueryID = dN["question"]["QueryID"]
                        OptionNDiagnosticID = dN["question"]["DiagnosticID"]
                        OptionNQuestion = dN["question"]["Question"]

                    elif "Request Confirmed" in isLastQuestion:
                        #print(optionSelected + " ---- " + isLastQuestion)
                        LastQuestionDiagnosticID = dN["question"]["DiagnosticID"]
                        count = 0
                        # ----------------Inserting data in Excel
                        DataList = ["Last Question DiagnosticID", LastQuestionDiagnosticID]
                        columns = 100
                        DataListCounter = 0
                        for c in range(1, columns + 1):
                            cellData = sheet.cell(row=pyIDCount + 2, column=c).value
                            if cellData is None or cellData == "None":
                                sheet.cell(row=pyIDCount + 2, column=c).value = DataList[DataListCounter]
                                DataListCounter = DataListCounter + 1
                                if DataListCounter == len(DataList):
                                    workbook.save(dataSheetPath)
                                    break
                        break
                    elif "Did this resolve the issue?" in isLastQuestion:
                        OptionN = "No"
                        OptionNQueryID = dN["pxResults"][0]["QueryID"]
                        OptionNDiagnosticID = dN["question"]["DiagnosticID"]
                        OptionNQuestion = dN["question"]["Question"]
                    else:
                        OptionN = dN["pxResults"][0]["Option"]
                        OptionNQueryID = dN["pxResults"][0]["QueryID"]
                        OptionNDiagnosticID = dN["question"]["DiagnosticID"]
                        OptionNQuestion = dN["question"]["Question"]

                    # ----------------Inserting data in Excel
                    DataList = [OptionNQuestion, OptionN]
                    columns = 100
                    DataListCounter = 0
                    for c in range(1, columns + 1):
                        cellData = sheet.cell(row=pyIDCount + 2, column=c).value
                        if cellData is None or cellData == "None":
                            sheet.cell(row=pyIDCount + 2, column=c).value = DataList[DataListCounter]
                            DataListCounter = DataListCounter + 1
                            if DataListCounter == len(DataList):
                                workbook.save(dataSheetPath)
                                break
                except:
                    #print("Here is the exception for no Question data in the response")
                    pass

            sheetName_Config = "jsonData"
            workbook = openpyxl.load_workbook(dataSheetPath)
            sheet = workbook[sheetName_Config]
            rows = sheet.max_row
            columns = sheet.max_column
            cellDataJson = ""
            for r in range(2, rows + 1):
                for c in range(1, columns + 1):
                    cellDataJson = sheet.cell(row=r, column=1).value
                    #print(cellDataJson)
            dJson = json.loads(cellDataJson)
            dJson["RequestorName"] = ResidentName
            dJson["TCNLocation"]["Address1"] = value[0]
            dJson["TCNLocation"]["City"] = value[1]
            dJson["TCNLocation"]["State"] = value[2]
            dJson["TCNLocation"]["ZipCode"] = value[3]
            dJson["MaintenanceRequests"][0]["DecisionTreeID"] = pyID
            #print("Request Data Sent - "+str(dJson))
            # ----------------post - maintenance request
            availableSlots = []
            headers = {"Authorization": "Bearer {}".format(token)}
            getD2API = '/1.0/maintenancerequest'
            r2 = requests.post(url=url + getD2API, headers=headers, json=dJson)
            d2 = json.loads(r2.text)
            #print(d2)
            Message = d2["ErrorMessages"][0]
            #print(Message)
            pyIDReceived = d2["WorkOrders"][0]["pyID"]
            #print(pyIDReceived)
            try:
                slotOriginReceived = d2["AvailableSlots"][0]["Origin"]
                print(slotOriginReceived)

                availableSlots.append(d2["AvailableSlots"][0]["FieldWorkerID"])
                availableSlots.append(d2["AvailableSlots"][0]["AppointmentStartDateTime"])
                availableSlots.append(d2["AvailableSlots"][0]["PreTravelTime"])
                availableSlots.append(d2["AvailableSlots"][0]["EndDateTime"])
                availableSlots.append(d2["AvailableSlots"][0]["SlotID"])
                availableSlots.append(d2["AvailableSlots"][0]["StartDateTime"])

                if "%26" in optionSelected:
                    optionSelected = optionSelected.replace(" %26 ", "&")
                print(Message + " - " + pyIDReceived + " for resident: " + ResidentName + " and category: " + optionSelected + ",  pyID :" + pyID)
                WOCounter = WOCounter + 1
                done = time.time()
                elapsed = done - start
                print("Work Order Count: "+str(WOCounter) + " , created in "+str(round(elapsed, 2)) + " seconds")

                #======================================================Finding slot for scheduling the request=========================================================
                print("First Slot Details : "+str(availableSlots[0]) + ", " + str(availableSlots[1]) + ", " + str(availableSlots[2]) + ", " + str(availableSlots[
                        3]) + ", " + str(availableSlots[4]) + ", " + str(availableSlots[5]))

                # ----------------Initiating scheduling request
                sheetName_Config = "jsonData"
                workbook = openpyxl.load_workbook(dataSheetPath)
                sheet = workbook[sheetName_Config]
                rows = sheet.max_row
                columns = sheet.max_column
                cellDataJsonSch = ""
                for r in range(2, rows + 1):
                    for c in range(1, columns + 1):
                        cellDataJsonSch = sheet.cell(row=r, column=2).value
                        # print(cellDataJsonSch)
                dJsonSch = json.loads(cellDataJsonSch)
                dJsonSch["pyID"] = pyIDReceived
                dJsonSch["Slot"]["FieldWorkerID"] = availableSlots[0]
                FieldWorkerID = availableSlots[0]
                dJsonSch["Slot"]["AppointmentStartDateTime"] = availableSlots[1]
                dJsonSch["Slot"]["PreTravelTime"] = availableSlots[2]
                dJsonSch["Slot"]["EndDateTime"] = availableSlots[3]
                dJsonSch["Slot"]["SlotID"] = availableSlots[4]
                dJsonSch["Slot"]["StartDateTime"] = availableSlots[5]
                #print("Request Data Sent - "+str(dJsonSch))
                # ----------------PUT - scheduling request
                headers = {"Authorization": "Bearer {}".format(token)}
                getD2API = '/1.0/schedulingrequest'
                r2 = requests.put(url=url + getD2API, headers=headers, json=dJsonSch)
                dJsonSchRes = json.loads(r2.text)
                Message = dJsonSchRes["ErrorMessages"][0]
                MessageCode = dJsonSchRes["ErrorCode"]
                doneFinal = time.time()
                elapsedFinal = doneFinal - start
                print("Response Received - " + Message + " , Response Code - " + str(MessageCode))
                if MessageCode == "200":
                    pass
                print("WO Created and Scheduled in "+str(round(elapsedFinal, 2)) + " seconds")

            except:
                WOCounter = WOCounter + 1
                done = time.time()
                elapsed = done - start
                FieldWorkerID = "None"
                elapsedFinal = elapsed
                print("No available slot found")
                print(pyIDReceived + " WO Created without scheduling and Work Order Count: " + str(WOCounter) + " , created in " + str(round(elapsed, 2)) + " seconds")

            # ----------------Saving Word Order data in Excel
            DateNow = datetime.datetime.now().strftime('%m-%d-%Y %H:%M:%S')
            sheetName_Config = "CreatedData"
            workbook = openpyxl.load_workbook(dataSheetPath)
            sheet = workbook[sheetName_Config]
            rowsMax = sheet.max_row
            columnsMax = sheet.max_column
            for r in range(rowsMax+1, rowsMax + 2):
                for c in range(1, columnsMax + 1):
                    if c == 1:
                        sheet.cell(row=r, column=c).value = r-1
                    if c == 2:
                        sheet.cell(row=r, column=c).value = pyIDReceived
                    if c == 3:
                        sheet.cell(row=r, column=c).value = pyID
                    if c == 4:
                        sheet.cell(row=r, column=c).value = ResidentName
                    if c == 5:
                        sheet.cell(row=r, column=c).value = optionSelected
                    if c == 6:
                        sheet.cell(row=r, column=c).value = FieldWorkerID
                    if c == 7:
                        sheet.cell(row=r, column=c).value = DateNow
                    if c == 8:
                        sheet.cell(row=r, column=c).value = round(elapsedFinal, 2)
            workbook.save(dataSheetPath)
        except:
            pass
    sheetName_Config = "PyIDList"
    workbook = openpyxl.load_workbook(dataSheetPath)
    sheet = workbook[sheetName_Config]
    rows = sheet.max_row
    columns = sheet.max_column
    for r in range(2, rows + 1):
        for c in range(1, columns + 1):
            sheet.cell(row=r, column=c).value = None
    workbook.save(dataSheetPath)
