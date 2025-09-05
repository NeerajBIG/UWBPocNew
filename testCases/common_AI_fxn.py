import os
import sys
from os.path import dirname, join, abspath

from testCases.AI_RandomData import Test_Data_Generator

sys.path.insert(0, abspath(join(dirname(__file__), '..')))
import pytesseract as pyt
import cv2
import openpyxl
from utilities.readProperties import ReadConfig

class ImageReader:
    basePath = ReadConfig.basePath()
    path1 = basePath
    text = path1.split("/")
    for t in text:
        if t == ".jenkins":
            print("Running on Jenkins")
            path = os.path.dirname(basePath)
            print("path is " + path)
            basePath = path
            JenkinsJobName = os.getenv("JOB_NAME")
            basePath = basePath + "/" + JenkinsJobName
        else:
            pass

    @staticmethod
    def imageReader(self, path):
        img = cv2.imread(path)
        pyt.pytesseract.tesseract_cmd = "C:\\Program Files\\Tesseract-OCR\\tesseract.exe"
        txt = pyt.image_to_string(img)
        words = txt.split()

        file_name = self.basePath+"\\TestData\\"+"ImageToWords_in_excel.xlsx"
        try:
            wb = openpyxl.load_workbook(file_name)
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active

        check_file = self.basePath+"\\TestData\\"+"ImageToWords_CompareText.xlsx"
        sheet_to_check = "ImageToWordLibrary"

        wb_check = openpyxl.load_workbook(check_file)
        sheet_check = wb_check[sheet_to_check]

        data_dict = {}
        for row in sheet_check.iter_rows(min_row=1, max_col=2, values_only=True):
            key = row[0]
            value = row[1]
            data_dict[key] = value

        words_to_check = list(data_dict.keys())

        if sheet.max_row == 1:
            empty_row = sheet.max_row
        else:
            empty_row = sheet.max_row + 1

        NoOfRecords = 2
        for col, word in enumerate(words, start=1):
            if word in words_to_check:
                sheet.cell(row=empty_row, column=col, value=word)
                #print("word: "+word + "   Type is: " + data_dict[word])

                for recordsCount in range(1, NoOfRecords + 1):
                    Value = Test_Data_Generator.test_generate_random_data_with_ai(self,data_dict[word])
                    sheet.cell(row=empty_row + recordsCount, column=col, value=Value)

        wb.save(file_name)
        print(f"Excel file saved as {file_name}")


